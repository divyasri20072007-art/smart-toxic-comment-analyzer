"""
utils.py
--------
Shared utility functions for ToxiShield AI:
  • EDA chart generation
  • WordCloud generation
  • Report export (CSV + XLSX)
  • Sample data generator (for demo mode)
"""

import io
import os
import random
import string

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from wordcloud import WordCloud

# ── Dark-theme defaults ───────────────────────────────────────────────────────
plt.rcParams.update({
    "figure.facecolor": "#0E1117",
    "axes.facecolor":   "#161B22",
    "axes.edgecolor":   "#30363D",
    "axes.labelcolor":  "#C9D1D9",
    "xtick.color":      "#8B949E",
    "ytick.color":      "#8B949E",
    "text.color":       "#C9D1D9",
    "grid.color":       "#21262D",
    "grid.linestyle":   "--",
    "grid.alpha":       0.5,
})

LABEL_COLS = ["toxic", "severe_toxic", "obscene", "threat", "insult", "identity_hate"]
COLORS     = ["#4ECDC4", "#FF6B6B", "#FFE66D", "#A8DADC", "#F4A261", "#E9C46A", "#2A9D8F"]


# ── Dataset statistics ────────────────────────────────────────────────────────
def dataset_stats(df: pd.DataFrame) -> dict:
    """Return a dictionary of high-level dataset statistics."""
    total      = len(df)
    toxic_mask = df[LABEL_COLS].any(axis=1)
    return {
        "Total Comments":     total,
        "Toxic Comments":     int(toxic_mask.sum()),
        "Non-Toxic Comments": int((~toxic_mask).sum()),
        "Missing Values":     int(df["comment_text"].isna().sum()),
        "Avg Comment Length": round(df["comment_text"].str.len().mean(), 1),
        "Max Comment Length": int(df["comment_text"].str.len().max()),
        "Min Comment Length": int(df["comment_text"].str.len().min()),
    }


# ── Class distribution chart ──────────────────────────────────────────────────
def plot_class_distribution(df: pd.DataFrame) -> plt.Figure:
    counts = {col.replace("_", " ").title(): int(df[col].sum()) for col in LABEL_COLS}
    counts["Non-Toxic"] = int((~df[LABEL_COLS].any(axis=1)).sum())

    labels = list(counts.keys())
    values = list(counts.values())
    cmap   = plt.cm.get_cmap("tab10", len(labels))
    colors = [cmap(i) for i in range(len(labels))]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    fig.patch.set_facecolor("#0E1117")

    # Bar chart
    ax1.set_facecolor("#161B22")
    bars = ax1.bar(labels, values, color=colors, edgecolor="#21262D", width=0.6)
    for bar, val in zip(bars, values):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 200,
                 f"{val:,}", ha="center", fontsize=9, color="#C9D1D9")
    ax1.set_title("Class Distribution (Count)", fontsize=13, fontweight="bold")
    ax1.set_xticklabels(labels, rotation=30, ha="right", fontsize=9)
    ax1.yaxis.grid(True)
    ax1.set_axisbelow(True)

    # Pie chart
    ax2.set_facecolor("#0E1117")
    wedges, texts, autotexts = ax2.pie(
        values, labels=labels, colors=colors,
        autopct="%1.1f%%", startangle=140,
        textprops={"color": "#C9D1D9", "fontsize": 8},
        wedgeprops={"edgecolor": "#0E1117", "linewidth": 1.5},
    )
    for at in autotexts:
        at.set_fontsize(8)
    ax2.set_title("Class Distribution (%)", fontsize=13, fontweight="bold")

    fig.tight_layout()
    return fig


# ── Comment length distribution ───────────────────────────────────────────────
def plot_comment_length_distribution(df: pd.DataFrame) -> plt.Figure:
    lengths = df["comment_text"].str.len().dropna()

    fig, ax = plt.subplots(figsize=(10, 4))
    fig.patch.set_facecolor("#0E1117")
    ax.set_facecolor("#161B22")

    ax.hist(lengths.clip(upper=2000), bins=60, color="#4ECDC4", edgecolor="#0E1117", alpha=0.85)
    ax.axvline(lengths.mean(), color="#FF6B6B", linestyle="--", linewidth=1.5,
               label=f"Mean: {lengths.mean():.0f}")
    ax.set_xlabel("Comment Length (chars)", fontsize=11)
    ax.set_ylabel("Frequency", fontsize=11)
    ax.set_title("Comment Length Distribution", fontsize=13, fontweight="bold")
    ax.legend(fontsize=10)
    ax.yaxis.grid(True)
    ax.set_axisbelow(True)
    fig.tight_layout()
    return fig


# ── Top-N word frequency bar chart ────────────────────────────────────────────
def plot_word_frequency(df: pd.DataFrame, text_col: str = "comment_text", top_n: int = 20) -> plt.Figure:
    from collections import Counter
    import re

    all_words = " ".join(df[text_col].dropna()).lower()
    words     = re.findall(r"\b[a-z]{3,}\b", all_words)
    common    = Counter(words).most_common(top_n)
    labels, counts = zip(*common) if common else ([], [])

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor("#0E1117")
    ax.set_facecolor("#161B22")

    bars = ax.barh(list(reversed(labels)), list(reversed(counts)),
                   color="#4ECDC4", edgecolor="#21262D", height=0.7)
    ax.set_xlabel("Frequency", fontsize=11)
    ax.set_title(f"Top {top_n} Most Frequent Words", fontsize=13, fontweight="bold")
    ax.xaxis.grid(True)
    ax.set_axisbelow(True)
    fig.tight_layout()
    return fig


# ── WordCloud generators ──────────────────────────────────────────────────────
def generate_wordcloud(text: str, title: str = "WordCloud",
                       colormap: str = "Reds") -> plt.Figure:
    wc = WordCloud(
        width=900, height=450,
        background_color="#0E1117",
        colormap=colormap,
        max_words=200,
        contour_width=1,
        contour_color="#30363D",
    ).generate(text or "no data available")

    fig, ax = plt.subplots(figsize=(10, 5))
    fig.patch.set_facecolor("#0E1117")
    ax.imshow(wc, interpolation="bilinear")
    ax.axis("off")
    ax.set_title(title, fontsize=14, fontweight="bold", color="#C9D1D9", pad=12)
    fig.tight_layout()
    return fig


def toxic_wordcloud(df: pd.DataFrame) -> plt.Figure:
    mask  = df[LABEL_COLS].any(axis=1)
    texts = " ".join(df.loc[mask, "comment_text"].dropna())
    return generate_wordcloud(texts, title="☁ WordCloud – Toxic Comments", colormap="Reds")


def nontoxic_wordcloud(df: pd.DataFrame) -> plt.Figure:
    mask  = ~df[LABEL_COLS].any(axis=1)
    texts = " ".join(df.loc[mask, "comment_text"].dropna().head(5000))
    return generate_wordcloud(texts, title="☁ WordCloud – Non-Toxic Comments", colormap="Greens")


# ── Excel / CSV export helpers ────────────────────────────────────────────────
def predictions_to_csv(df: pd.DataFrame) -> bytes:
    """Serialise predictions DataFrame to CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


def predictions_to_xlsx(df: pd.DataFrame) -> bytes:
    """Serialise predictions DataFrame to XLSX bytes with styling."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Predictions")

        wb = writer.book
        ws = writer.sheets["Predictions"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin", color="374151"),
            right=Side(style="thin", color="374151"),
            top=Side(style="thin", color="374151"),
            bottom=Side(style="thin", color="374151"),
        )
        risk_colors = {
            "Critical": "FF4C4C",
            "High":     "FF8C42",
            "Medium":   "FFD166",
            "Low":      "06D6A0",
        }

        # Style header row
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border

        # Style data rows
        risk_col_idx = None
        for idx, col in enumerate(df.columns, 1):
            if col == "risk_level":
                risk_col_idx = idx

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if risk_col_idx and row[risk_col_idx - 1].value in risk_colors:
                color = risk_colors[row[risk_col_idx - 1].value]
                row[risk_col_idx - 1].fill = PatternFill("solid", fgColor=color)
                row[risk_col_idx - 1].font = Font(bold=True, color="000000")

        # Auto column widths
        for i, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 60)

        ws.row_dimensions[1].height = 20

        # Summary sheet
        summary_data = df["predicted_label"].value_counts().reset_index()
        summary_data.columns = ["Category", "Count"]
        summary_data.to_excel(writer, index=False, sheet_name="Summary")

    return buf.getvalue()


# ── Demo / sample data ────────────────────────────────────────────────────────
SAMPLE_COMMENTS = [
    "I love this community! Everyone is so helpful and kind.",
    "You are a complete idiot and should never post again.",
    "Great article, really informative. Keep up the good work!",
    "I will find you and hurt you. You better watch out.",
    "This is the dumbest thing I have ever seen, you moron.",
    "Thanks for sharing this wonderful tutorial!",
    "People like you should be banned from the internet forever.",
    "What a fascinating perspective, I had never thought of it that way.",
    "Go back to your country, nobody wants you here.",
    "The weather today is absolutely beautiful!",
]


def generate_sample_df(n: int = 10) -> pd.DataFrame:
    """Return a small demo DataFrame for batch prediction testing."""
    comments = random.choices(SAMPLE_COMMENTS, k=n)
    return pd.DataFrame({"comment_text": comments})
